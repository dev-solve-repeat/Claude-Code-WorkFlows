"""
Tool: export_to_excel.py
Purpose: Convert a scraped jobs JSON file into a formatted Excel spreadsheet.
Output: Job Listings - {search} ({date}).xlsx in the project root

Usage:
    python tools/export_to_excel.py .tmp/jobs_inbound_sales.json
    python tools/export_to_excel.py .tmp/jobs_inbound_sales.json --output "My Jobs.xlsx"
"""

import argparse
import json
import os
import sys
from datetime import datetime

try:
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
except ImportError:
    print("[error] openpyxl is not installed. Run: pip3 install openpyxl")
    sys.exit(1)


COLUMNS = [
    ("Title",       "title",                50),
    ("Job Type",    "job_type",             14),
    ("Location",    "location",             20),
    ("Salary",      "salary",               25),
    ("Experience",  "experience",           16),
    ("Category",    "category",             18),
    ("Role",        "role",                 28),
    ("Post Date",   "post_date",            14),
    ("Description", "description_snippet",  65),
    ("Job URL",     "job_url",              55),
]

HEADER_FILL  = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_FONT  = Font(color="FFFFFF", bold=True, size=11)
HEADER_ALIGN = Alignment(horizontal="center", vertical="center")

ALT_FILL_EVEN = PatternFill(start_color="EBF3FB", end_color="EBF3FB", fill_type="solid")
ALT_FILL_ODD  = PatternFill(fill_type=None)  # white / no fill

URL_FONT = Font(color="0563C1", underline="single", size=10)

DESC_COL_IDX = next(i + 1 for i, (_, key, _) in enumerate(COLUMNS) if key == "description_snippet")
URL_COL_IDX  = next(i + 1 for i, (_, key, _) in enumerate(COLUMNS) if key == "job_url")


def apply_header(ws):
    for col_idx, (header, _, _) in enumerate(COLUMNS, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill  = HEADER_FILL
        cell.font  = HEADER_FONT
        cell.alignment = HEADER_ALIGN
    ws.row_dimensions[1].height = 22


def set_column_widths(ws):
    for col_idx, (_, _, width) in enumerate(COLUMNS, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width


def write_jobs(ws, jobs: list):
    for row_offset, job in enumerate(jobs):
        row_num = row_offset + 2  # data starts at row 2
        fill = ALT_FILL_EVEN if row_num % 2 == 0 else ALT_FILL_ODD

        for col_idx, (_, field, _) in enumerate(COLUMNS, 1):
            value = job.get(field, "")
            cell = ws.cell(row=row_num, column=col_idx, value=value)
            cell.fill = fill

            # Description column: wrap text
            if col_idx == DESC_COL_IDX:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
            else:
                cell.alignment = Alignment(vertical="top", wrap_text=False)

            # URL column: hyperlink + blue underline
            if col_idx == URL_COL_IDX and value:
                cell.hyperlink = value
                cell.font = URL_FONT


def main():
    parser = argparse.ArgumentParser(
        description="Export scraped jobs JSON to a formatted Excel file"
    )
    parser.add_argument("input_path", help="Path to .tmp/jobs_{slug}.json")
    parser.add_argument("--output", default=None, help="Override output .xlsx path")
    args = parser.parse_args()

    # Load JSON
    if not os.path.exists(args.input_path):
        print(f"[error] Input file not found: {args.input_path}")
        sys.exit(1)

    with open(args.input_path, "r", encoding="utf-8") as f:
        data = json.load(f)

    jobs = data.get("jobs", [])
    search = data.get("search", "Jobs")

    if not jobs:
        print(f"[warn] No jobs found in {args.input_path} — creating empty Excel with headers only")

    # Derive output path
    if args.output:
        output_path = args.output
    else:
        date_str = datetime.now().strftime("%Y-%m-%d")
        filename = f"Job Listings - {search} ({date_str}).xlsx"
        # Place in project root (one level up from tools/)
        project_root = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
        output_path = os.path.join(project_root, filename)

    # Build workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Job Listings"
    wb.properties.title = f"Job Listings - {search}"
    wb.properties.creator = "WAT Framework / scrape_jobs"

    apply_header(ws)
    set_column_widths(ws)
    write_jobs(ws, jobs)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    # Ensure output directory exists
    out_dir = os.path.dirname(output_path)
    if out_dir:
        os.makedirs(out_dir, exist_ok=True)

    wb.save(output_path)

    print(f"\n--- Export Complete ---")
    print(f"  Rows written: {len(jobs)}")
    print(f"  Output:       {output_path}")


if __name__ == "__main__":
    main()
