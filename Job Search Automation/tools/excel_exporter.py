"""
excel_exporter.py — Export scraped jobs to a formatted Excel workbook.

Reads all .tmp/jobs_raw_*.json files produced by job_scraper.py.
Output: output/job_matches_YYYY-MM-DD.xlsx
"""

import json
import sys
from datetime import datetime
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("ERROR: openpyxl not installed. Run: pip install openpyxl")
    sys.exit(1)

ROOT = Path(__file__).parent.parent
TMP_DIR = ROOT / ".tmp"
OUTPUT_DIR = ROOT / "output"

COLUMNS = [
    ("Company",              28),
    ("Job Role",             35),
    ("Matched Roles",        38),
    ("Yrs Exp Required",     16),
    ("Job Description",      55),
    ("Job Requirements",     55),
    ("Skills Required",      38),
    ("Location",             22),
    ("Remote?",              10),
    ("Portal",               12),
    ("Job URL",              45),
    ("Posted Date",          18),
    ("Date Scraped",         18),
]

HEADER_FILL   = PatternFill("solid", fgColor="1F3864")
ROW_FILL_ODD  = PatternFill("solid", fgColor="EEF2FF")
ROW_FILL_EVEN = PatternFill("solid", fgColor="FFFFFF")
HEADER_FONT   = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
DATA_FONT     = Font(name="Calibri", size=10)
THIN_BORDER   = Border(
    left=Side(style="thin", color="D0D0D0"),
    right=Side(style="thin", color="D0D0D0"),
    top=Side(style="thin", color="D0D0D0"),
    bottom=Side(style="thin", color="D0D0D0"),
)


def load_jobs() -> list[dict]:
    raw_files = sorted(TMP_DIR.glob("jobs_raw_*.json"))
    if not raw_files:
        print("ERROR: No job data found in .tmp/")
        print("Run tools/job_scraper.py first.")
        sys.exit(1)

    all_jobs: list[dict] = []
    for f in raw_files:
        try:
            jobs = json.loads(f.read_text())
            if jobs:
                all_jobs.extend(jobs)
                print(f"  Loaded {len(jobs):>4} jobs from {f.name}")
        except json.JSONDecodeError:
            print(f"  WARNING: Could not parse {f.name} — skipping")

    return all_jobs


def build_row(job: dict) -> list:
    matched_roles = job.get("matched_roles", [])
    matched_str = ", ".join(matched_roles) if isinstance(matched_roles, list) else str(matched_roles)
    remote_val = "Yes" if job.get("remote") else "No"

    # Normalise posted_at to date-only string
    posted_raw = job.get("posted_at", "")
    if posted_raw and "T" in posted_raw:
        posted_raw = posted_raw[:10]

    scraped_raw = job.get("scraped_at", "")
    if scraped_raw and "T" in scraped_raw:
        scraped_raw = scraped_raw[:10]

    return [
        job.get("company", ""),
        job.get("title", ""),
        matched_str,
        job.get("exp_required", ""),
        job.get("description", ""),
        job.get("requirements", ""),
        job.get("skills", ""),
        job.get("location", ""),
        remote_val,
        job.get("portal", ""),
        job.get("url", ""),
        posted_raw,
        scraped_raw,
    ]


def apply_header(ws):
    for col_idx, (col_name, col_width) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = col_width
    ws.row_dimensions[1].height = 30


def apply_data_row(ws, row_idx: int, values: list, url_col_idx: int):
    row_fill = ROW_FILL_ODD if row_idx % 2 == 0 else ROW_FILL_EVEN

    for col_idx, value in enumerate(values, start=1):
        cell = ws.cell(row=row_idx, column=col_idx, value=value or "")
        cell.border = THIN_BORDER
        cell.alignment = Alignment(vertical="top", wrap_text=True)
        cell.fill = row_fill

        if col_idx == url_col_idx and value:
            cell.hyperlink = str(value)
            cell.font = Font(name="Calibri", size=10, color="0563C1", underline="single")
        else:
            cell.font = DATA_FONT

    ws.row_dimensions[row_idx].height = 60


def create_summary_sheet(wb: Workbook, jobs: list[dict]):
    ws = wb.create_sheet("Summary", 0)
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 36
    ws.column_dimensions["B"].width = 20

    portals: dict[str, int] = {}
    role_counts: dict[str, int] = {}
    remote_count = sum(1 for j in jobs if j.get("remote"))

    for j in jobs:
        p = j.get("portal", "unknown")
        portals[p] = portals.get(p, 0) + 1
        for role in j.get("matched_roles", []):
            role_counts[role] = role_counts.get(role, 0) + 1

    rows: list[tuple] = [
        ("Job Search Summary", ""),
        ("", ""),
        ("Run Date", datetime.now().strftime("%Y-%m-%d %H:%M")),
        ("Total Jobs", len(jobs)),
        ("Remote Jobs", remote_count),
        ("India Jobs (non-remote)", sum(1 for j in jobs if j.get("portal") == "naukri" and not j.get("remote"))),
        ("", ""),
        ("Jobs by Portal", ""),
    ]
    for portal, count in sorted(portals.items(), key=lambda x: -x[1]):
        rows.append((f"  {portal.capitalize()}", count))

    rows += [("", ""), ("Jobs by Matched Role", "")]
    for role, count in sorted(role_counts.items(), key=lambda x: -x[1]):
        rows.append((f"  {role}", count))

    for r_idx, (label, value) in enumerate(rows, start=1):
        cell_a = ws.cell(row=r_idx, column=1, value=label)
        cell_b = ws.cell(row=r_idx, column=2, value=value)

        if label in ("Job Search Summary", "Jobs by Portal", "Jobs by Matched Role"):
            cell_a.font = Font(name="Calibri", bold=True, size=13)
        else:
            cell_a.font = Font(name="Calibri", size=11)
            cell_b.font = Font(name="Calibri", size=11, bold=bool(value))


def main():
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    print("Loading jobs...")
    jobs = load_jobs()

    if not jobs:
        print("No jobs to export.")
        sys.exit(0)

    print(f"\nExporting {len(jobs)} jobs to Excel...")

    wb = Workbook()
    ws = wb.active
    ws.title = "All Jobs"
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A2"

    url_col_idx = next(i + 1 for i, (name, _) in enumerate(COLUMNS) if name == "Job URL")

    apply_header(ws)

    for row_idx, job in enumerate(jobs, start=2):
        apply_data_row(ws, row_idx, build_row(job), url_col_idx)

    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}1"

    create_summary_sheet(wb, jobs)
    wb.active = wb["Summary"]

    today = datetime.now().strftime("%Y-%m-%d")
    out_path = OUTPUT_DIR / f"job_matches_{today}.xlsx"
    wb.save(out_path)

    portals = len(set(j.get("portal", "") for j in jobs))
    print(f"\nExcel saved: {out_path}")
    print(f"  {len(jobs)} jobs | {portals} portals")


if __name__ == "__main__":
    main()
